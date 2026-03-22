"""
routes/documents.py - загрузка, удаление, запросы к документам, отчеты OCR
"""

import logging
import os
from datetime import datetime

from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import FileResponse

from backend.app_state import ask_agent, rag_client, minio_client, settings
from backend.schemas import DocumentQueryRequest
from backend.socket_helpers import _is_structure_query

router = APIRouter(prefix="/api/documents", tags=["documents"])
logger = logging.getLogger(__name__)


@router.post("/upload")
async def upload_document(file: UploadFile = File(...)):
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")

    documents_bucket = os.getenv("MINIO_DOCUMENTS_BUCKET_NAME", "astrachat-documents")
    file_object_name = None

    try:
        content = await file.read()
        file_extension = os.path.splitext(file.filename)[1].lower() if file.filename else ""
        is_image = file_extension in (".jpg", ".jpeg", ".png", ".webp")
        content_type_map = {
            ".pdf": "application/pdf", ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ".doc": "application/msword", ".txt": "text/plain",
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".xls": "application/vnd.ms-excel", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
            ".png": "image/png", ".webp": "image/webp",
        }
        content_type = content_type_map.get(file_extension, "application/octet-stream")

        if minio_client:
            try:
                file_object_name = minio_client.generate_object_name(
                    prefix="img_" if is_image else "doc_", extension=file_extension
                )
                minio_client.upload_file(content, file_object_name, content_type=content_type, bucket_name=documents_bucket)
            except Exception as e:
                logger.warning(f"MinIO upload: {e}")
                file_object_name = None

        try:
            rag_result = await rag_client.upload_document(
                file_bytes=content,
                filename=file.filename or file_object_name or "unknown",
                minio_object=file_object_name,
                minio_bucket=documents_bucket if minio_client and file_object_name else None,
                original_path=None,
            )
        except Exception as e:
            if minio_client and file_object_name:
                try:
                    minio_client.delete_file(file_object_name, bucket_name=documents_bucket)
                except Exception:
                    pass
            raise HTTPException(status_code=502, detail=f"Ошибка RAG-сервиса: {e}")

        if not rag_result.get("ok"):
            if minio_client and file_object_name:
                try:
                    minio_client.delete_file(file_object_name, bucket_name=documents_bucket)
                except Exception:
                    pass
            raise HTTPException(status_code=400, detail=rag_result.get("error", "Ошибка индексации"))

        result = {"message": "Документ успешно загружен", "filename": file.filename,
                  "success": True, "rag_document_id": rag_result.get("document_id")}
        if is_image and minio_client and file_object_name:
            result["minio_object"] = file_object_name
            result["minio_bucket"] = documents_bucket
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/query")
async def query_document(request: DocumentQueryRequest):
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    if not ask_agent:
        raise HTTPException(status_code=503, detail="AI agent не доступен")
    try:
        from backend.app_state import current_rag_strategy
        hits = await rag_client.search(request.query, k=12, strategy=current_rag_strategy)
        if not hits:
            return {"response": "В загруженных документах не найдено информации.", "query": request.query, "success": True}

        if _is_structure_query(request.query):
            seen = {(d, i) for _, _, d, i in hits}
            for doc_id in {d for _, _, d, _ in hits if d}:
                try:
                    for c, sc, did, idx in await rag_client.get_document_start_chunks(doc_id, max_chunks=2):
                        if (did, idx) not in seen:
                            hits = [(c, sc, did, idx)] + hits
                            seen.add((did, idx))
                except Exception:
                    pass

        parts, total = [], 0
        for i, (content, score, doc_id, chunk_idx) in enumerate(hits, 1):
            frag = f"Фрагмент {i} (document_id={doc_id}, чанк {chunk_idx}, релевантность: {score:.2f}):\n{content}\n"
            if total + len(frag) > 12000:
                parts.append(frag[:max(0, 12000 - total - 80)] + "\n... [обрезано]\n")
                break
            parts.append(frag)
            total += len(frag)

        prompt = (
            f"На основе контекста из документов ответь на вопрос. Не придумывай информацию.\n\n"
            f"Контекст:\n{chr(10).join(parts)}\n\nВопрос: {request.query}\n\nОтвет:"
        )
        response_text = ask_agent(prompt)
        return {"response": response_text, "query": request.query, "success": True, "timestamp": datetime.now().isoformat()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("")
@router.get("/")
async def get_documents():
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    try:
        docs = await rag_client.list_documents()
        filenames = [d.get("filename") for d in docs]
        return {"documents": filenames, "count": len(filenames), "success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{filename}")
async def delete_document(filename: str):
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    try:
        docs = await rag_client.list_documents()
        filenames = [d.get("filename") for d in docs]
        if filename not in filenames:
            raise HTTPException(status_code=404, detail=f"Документ {filename} не найден")

        documents_bucket = os.getenv("MINIO_DOCUMENTS_BUCKET_NAME", "astrachat-documents")
        if minio_client:
            try:
                minio_info = await rag_client.get_image_minio_info(filename)
                if minio_info:
                    minio_client.delete_file(minio_info["minio_object"], bucket_name=minio_info["minio_bucket"])
            except Exception as e:
                logger.warning(f"MinIO delete: {e}")

        try:
            await rag_client.delete_document_by_filename(filename)
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Ошибка RAG-сервиса: {e}")

        new_docs = await rag_client.list_documents()
        return {"message": f"Документ {filename} удален", "success": True,
                "remaining_documents": [d.get("filename") for d in new_docs]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/report/generate")
async def generate_confidence_report():
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    try:
        report_data = await rag_client.get_confidence_report()
        logger.info(f"Получены данные отчета: {report_data['total_documents']} документов")

        report_text = f"""
        ОТЧЕТ О СТЕПЕНИ УВЕРЕННОСТИ МОДЕЛИ В РАСПОЗНАННОМ ТЕКСТЕ
        {'=' * 80}
        Дата генерации: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        {'=' * 80}
        
        ОБЩАЯ ИНФОРМАЦИЯ:
        - Всего обработано документов: {report_data['total_documents']}
        - Средняя уверенность модели: {report_data['average_confidence']:.2f}%
        - Всего слов: {report_data.get('total_words', 0)}
        {'=' * 80}
        ДЕТАЛЬНАЯ ИНФОРМАЦИЯ ПО ДОКУМЕНТАМ:
        """

        for i, doc in enumerate(report_data['documents'], 1):
            report_text += f"""

{i}. {doc['filename']}
   Тип файла: {doc['file_type']}
   Уверенность модели: {doc['confidence']:.2f}%
   Длина распознанного текста: {doc['text_length']} символов
   Количество слов: {doc.get('words_count', 0)}
   {'-' * 80}
   
   РАСПОЗНАННЫЙ ТЕКСТ С УВЕРЕННОСТЬЮ:
"""
            formatted_text_info = next(
                (ft for ft in report_data.get('formatted_texts', []) if ft['filename'] == doc['filename']), None
            )

            if formatted_text_info and formatted_text_info.get('words'):
                words = formatted_text_info.get('words', [])
                if words:
                    line_words = []
                    current_line = []
                    for word_info in words:
                        word = word_info.get('word', '')
                        conf = word_info.get('confidence', 0.0)
                        current_line.append((word, conf))
                        if len(current_line) >= 8:
                            line_words.append(current_line)
                            current_line = []
                    if current_line:
                        line_words.append(current_line)

                    if line_words:
                        for line in line_words:
                            import re
                            tokens_data = []
                            prev_is_punctuation = False
                            for word, conf in line:
                                is_punctuation = bool(re.match(r'^[^\w\s]+$', word))
                                word_width = len(word)
                                col_width = max(word_width + 2, 10)
                                tokens_data.append({
                                    'word': word, 'conf': conf,
                                    'is_punctuation': is_punctuation, 'col_width': col_width,
                                    'needs_space_before': not prev_is_punctuation and not is_punctuation and tokens_data,
                                })
                                prev_is_punctuation = is_punctuation

                            percent_line = "│"
                            word_line = "│"
                            separator_line = "├"
                            for idx, token in enumerate(tokens_data):
                                if token['needs_space_before']:
                                    word_line += "│"
                                    percent_line += "│"
                                    separator_line += "┼"
                                percent_str = f"{token['conf']:.0f}%"
                                word_str = token['word']
                                percent_padded = percent_str.center(token['col_width'])
                                word_padded = word_str.ljust(token['col_width'])
                                percent_line += percent_padded + "│"
                                word_line += word_padded + "│"
                                separator_line += "─" * token['col_width'] + (
                                    "┤" if idx == len(tokens_data) - 1 else "┼"
                                )
                            report_text += f"{percent_line}\n"
                            report_text += f"{separator_line}\n"
                            report_text += f"{word_line}\n\n"
                    else:
                        report_text += "[Нет валидных слов для отображения]\n"
                else:
                    report_text += "[Нет данных о словах]\n"
            else:
                report_text += "[Нет отформатированного текста]\n"

            report_text += f"{'-' * 80}\n"

        overall_conf = report_data.get('overall_confidence', report_data.get('average_confidence', 0.0))
        report_text += f"""

{'=' * 80}
ИТОГО:
- Итоговая уверенность по всему распознанному тексту: {overall_conf:.2f}%
- Средняя уверенность по документам: {report_data['average_confidence']:.2f}%
- Всего документов: {report_data['total_documents']}
- Всего слов: {report_data.get('total_words', 0)}
{'=' * 80}
"""

        report_json = {
            "generated_at": datetime.now().isoformat(),
            "summary": {
                "total_documents": report_data['total_documents'],
                "average_confidence": round(report_data['average_confidence'], 2),
                "overall_confidence": round(overall_conf, 2),
                "total_words": report_data.get('total_words', 0),
            },
            "documents": report_data['documents'],
        }

        return {
            "success": True,
            "report_text": report_text,
            "report_json": report_json,
            "summary": {
                "total_documents": report_data['total_documents'],
                "average_confidence": round(report_data['average_confidence'], 2),
                "overall_confidence": round(overall_conf, 2),
                "total_words": report_data.get('total_words', 0),
            },
        }
    except Exception as e:
        logger.error(f"Ошибка при генерации отчета: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/report/download")
async def download_confidence_report():
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import tempfile

        report_data = await rag_client.get_confidence_report()
        logger.info(f"Получены данные отчета: {report_data['total_documents']} документов")

        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет об уверенности"

        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        subheader_font = Font(bold=True, size=12)
        subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        high_confidence_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        medium_confidence_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        low_confidence_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        current_row = 1

        ws.merge_cells(f'A{current_row}:D{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = "ОТЧЕТ О СТЕПЕНИ УВЕРЕННОСТИ МОДЕЛИ В РАСПОЗНАННОМ ТЕКСТЕ"
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment
        header_cell.border = thin_border
        current_row += 1

        ws.merge_cells(f'A{current_row}:D{current_row}')
        date_cell = ws[f'A{current_row}']
        date_cell.value = f"Дата генерации: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.alignment = Alignment(horizontal="center")
        current_row += 2

        if report_data['total_documents'] == 0:
            ws.merge_cells(f'A{current_row}:D{current_row}')
            warning_cell = ws[f'A{current_row}']
            warning_cell.value = "ВНИМАНИЕ: Нет обработанных документов для формирования отчета."
            warning_cell.font = Font(bold=True, color="FF0000")
            warning_cell.alignment = Alignment(horizontal="center")
            current_row += 1
        else:
            info_row = current_row
            ws[f'A{info_row}'] = "ОБЩАЯ ИНФОРМАЦИЯ:"
            ws[f'A{info_row}'].font = subheader_font
            ws[f'A{info_row}'].fill = subheader_fill
            current_row += 1

            ws[f'A{current_row}'] = "Всего обработано документов:"
            ws[f'B{current_row}'] = report_data['total_documents']
            current_row += 1

            ws[f'A{current_row}'] = "Средняя уверенность модели:"
            ws[f'B{current_row}'] = f"{report_data['average_confidence']:.2f}%"
            current_row += 1

            ws[f'A{current_row}'] = "Всего слов:"
            ws[f'B{current_row}'] = report_data.get('total_words', 0)
            current_row += 2

            for doc_idx, doc in enumerate(report_data.get('documents', []), 1):
                ws.merge_cells(f'A{current_row}:D{current_row}')
                doc_header = ws[f'A{current_row}']
                doc_header.value = f"{doc_idx}. {doc.get('filename', 'Неизвестный файл')}"
                doc_header.font = subheader_font
                doc_header.fill = subheader_fill
                doc_header.border = thin_border
                current_row += 1

                ws[f'A{current_row}'] = "Тип файла:"
                ws[f'B{current_row}'] = doc.get('file_type', 'unknown')
                current_row += 1

                ws[f'A{current_row}'] = "Уверенность модели:"
                conf_value = doc.get('confidence', 0.0)
                ws[f'B{current_row}'] = f"{conf_value:.2f}%"
                if conf_value >= 80:
                    ws[f'B{current_row}'].fill = high_confidence_fill
                elif conf_value >= 50:
                    ws[f'B{current_row}'].fill = medium_confidence_fill
                else:
                    ws[f'B{current_row}'].fill = low_confidence_fill
                current_row += 1

                ws[f'A{current_row}'] = "Длина текста:"
                ws[f'B{current_row}'] = f"{doc.get('text_length', 0)} символов"
                current_row += 1

                ws[f'A{current_row}'] = "Количество слов:"
                ws[f'B{current_row}'] = doc.get('words_count', 0)
                current_row += 2

                formatted_text_info = next(
                    (ft for ft in report_data.get('formatted_texts', []) if ft.get('filename') == doc.get('filename')),
                    None,
                )

                if formatted_text_info and formatted_text_info.get('words'):
                    words = formatted_text_info.get('words', [])
                    if words:
                        ws[f'A{current_row}'] = "Слово"
                        ws[f'B{current_row}'] = "Уверенность"
                        ws[f'A{current_row}'].font = Font(bold=True)
                        ws[f'B{current_row}'].font = Font(bold=True)
                        ws[f'A{current_row}'].fill = subheader_fill
                        ws[f'B{current_row}'].fill = subheader_fill
                        ws[f'A{current_row}'].border = thin_border
                        ws[f'B{current_row}'].border = thin_border
                        current_row += 1

                        for word_info in words:
                            word = word_info.get('word', '')
                            conf = word_info.get('confidence', 0.0)
                            if word:
                                ws[f'A{current_row}'] = word
                                ws[f'B{current_row}'] = f"{conf:.1f}%"
                                ws[f'A{current_row}'].border = thin_border
                                ws[f'B{current_row}'].border = thin_border
                                if conf >= 80:
                                    ws[f'B{current_row}'].fill = high_confidence_fill
                                elif conf >= 50:
                                    ws[f'B{current_row}'].fill = medium_confidence_fill
                                else:
                                    ws[f'B{current_row}'].fill = low_confidence_fill
                                current_row += 1
                current_row += 1

            overall_conf = report_data.get('overall_confidence', report_data.get('average_confidence', 0.0))
            ws.merge_cells(f'A{current_row}:D{current_row}')
            summary_header = ws[f'A{current_row}']
            summary_header.value = "ИТОГО"
            summary_header.font = subheader_font
            summary_header.fill = subheader_fill
            summary_header.border = thin_border
            current_row += 1

            ws[f'A{current_row}'] = "Итоговая уверенность по всему тексту:"
            ws[f'B{current_row}'] = f"{overall_conf:.2f}%"
            if overall_conf >= 80:
                ws[f'B{current_row}'].fill = high_confidence_fill
            elif overall_conf >= 50:
                ws[f'B{current_row}'].fill = medium_confidence_fill
            else:
                ws[f'B{current_row}'].fill = low_confidence_fill
            current_row += 1

            ws[f'A{current_row}'] = "Средняя уверенность по документам:"
            ws[f'B{current_row}'] = f"{report_data['average_confidence']:.2f}%"
            current_row += 1

            ws[f'A{current_row}'] = "Всего документов:"
            ws[f'B{current_row}'] = report_data['total_documents']
            current_row += 1

            ws[f'A{current_row}'] = "Всего слов:"
            ws[f'B{current_row}'] = report_data.get('total_words', 0)

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        temp_dir = tempfile.gettempdir()
        report_filename = f"confidence_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = os.path.join(temp_dir, report_filename)

        try:
            os.makedirs(temp_dir, exist_ok=True)
            wb.save(report_path)
            logger.info(f"Excel отчет сохранен: {report_path}")
            return FileResponse(
                report_path,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename=report_filename,
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{report_filename}"},
            )
        except Exception as file_err:
            logger.error(f"Ошибка при сохранении Excel: {file_err}")
            raise HTTPException(status_code=500, detail=str(file_err))

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при генерации Excel: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))
